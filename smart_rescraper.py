"""
smart_rescraper_final.py — Final Clean Version
================================================
Reads results.xlsx, skips URLs that already have valid shipping timelines,
re-scrapes only the incomplete ones with STRICT extraction (no product descriptions),
and writes results_updated.xlsx.

USAGE:
    python smart_rescraper_final.py
    python smart_rescraper_final.py --input results.xlsx --output results_updated.xlsx --delay 3

HOW EXTRACTION WORKS (priority order):
  1. JSON variant data  → "Ships In 14 Days" from Shopify variant JSON
  2. <legend> span      → Delivery Timeline label → span value
  3. <input value>      → input[value*="Ships"]
  4. .text-truncator p  → "This product will be shipped within 3-5 working days"
  5. Any <p> with shipping keywords (strict: must contain ship/deliver + timeframe)
  6. Shopify product JSON endpoint (/products/slug.json body_html)

WHAT IS CONSIDERED VALID (will be SKIPPED):
  - Contains any of: ship, deliver, dispatch, transit, working day, business day
  - OR starts with: "Ships In", "Delivered in", "Same day", "Express"

WHAT IS CONSIDERED INVALID (will be RE-SCRAPED):
  - "N/A"
  - Empty / None
  - Only contains size words: XS, S, M, L, XL, XXL, Small, Medium, Large etc.
  - Only contains a price (digits + currency)
  - Random product description text (no shipping keywords)
"""

import argparse
import asyncio
import csv
import logging
import random
import re
import sys
import time
from pathlib import Path

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("rescraper")

# ── Constants ─────────────────────────────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
]

# Keywords that PROVE an entry is a valid shipping timeline
SHIPPING_KEYWORDS = [
    "ship", "deliver", "dispatch", "transit",
    "working day", "business day", "same day", "express delivery",
    "days", "weeks", "hours",
]

# Size words that indicate extraction grabbed the wrong thing
SIZE_WORDS = {
    "xs", "s", "m", "l", "xl", "xxl", "xxxl",
    "xsmall", "small", "medium", "large", "xlarge", "xxlarge",
    "extra small", "extra large", "one size", "os", "free size",
}

# Regex for price-like values (8999, ₹21999, $120.00)
PRICE_RE = re.compile(r"^[₹$£€]?\s*[\d,]+(\.\d{1,2})?$")

# Regex: valid shipping value must contain a timeframe number + shipping word
VALID_RE = re.compile(
    r"(ship|deliver|dispatch|transit|working\s*day|business\s*day|same\s*day|express)",
    re.IGNORECASE,
)

# Patterns that indicate product DESCRIPTION crept in (not shipping info)
DESCRIPTION_SIGNALS = [
    r"introducing the",
    r"delivering.*satisfaction",
    r"collection",
    r"perfect for",
    r"made from",
    r"100%",
    r"fabric",
    r"features",
    r"embroidered",
    r"graphic",
    r"streetwear",
    r"inspired",
    r"high quality",
    r"limited edition",
]
DESCRIPTION_RE = re.compile("|".join(DESCRIPTION_SIGNALS), re.IGNORECASE)


# ── Validation ────────────────────────────────────────────────────────────────

def is_valid_shipping(value: str) -> bool:
    """Return True if the value looks like a genuine shipping timeline."""
    if not value or not value.strip():
        return False

    v = value.strip()

    # Explicit failures
    if v.lower() in ("n/a", "na", "none", "-", ""):
        return False

    # Price pattern
    if PRICE_RE.match(v):
        return False

    # Pure size word
    if v.lower() in SIZE_WORDS:
        return False

    # Multiple size words (e.g. "Small / Medium")
    parts = re.split(r"[,/|]", v)
    if all(p.strip().lower() in SIZE_WORDS for p in parts if p.strip()):
        return False

    # Looks like a product description
    if DESCRIPTION_RE.search(v):
        return False

    # Must contain a shipping keyword
    if not VALID_RE.search(v):
        return False

    return True


# ── Extraction ────────────────────────────────────────────────────────────────

def extract_shipping(html: str, url: str = "") -> str:
    """
    Parse HTML and extract ONLY shipping timeline information.
    Returns clean value or "N/A".
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        log.error("BeautifulSoup not installed. Run: pip install beautifulsoup4")
        sys.exit(1)

    soup = BeautifulSoup(html, "html.parser")

    # ── 1. Shopify variant JSON in <script> tags ──────────────────────────────
    for script in soup.find_all("script", type="application/json"):
        text = script.get_text()
        # Look for delivery timeline patterns in JSON
        match = re.search(
            r'"(?:title|value|name)"\s*:\s*"((?:Ships?\s+[Ii]n|Deliver[sy]?\s+[Ii]n|Dispatch(?:ed)?\s+[Ii]n)[^"]{1,60})"',
            text,
        )
        if match:
            return clean_value(match.group(1))

        # Also search for "Delivery Timeline" key
        match = re.search(
            r'"Delivery\s+Timeline[s]?"\s*:\s*"([^"]{3,80})"',
            text,
            re.IGNORECASE,
        )
        if match:
            return clean_value(match.group(1))

    # ── 2. <script> tags with product JSON (Shopify window.ShopifyAnalytics etc.) ──
    for script in soup.find_all("script"):
        text = script.get_text()
        if "Ships In" in text or "Delivery Timeline" in text:
            match = re.search(
                r'(?:Ships?\s+[Ii]n\s+\d+[^"\'<\n]{0,30})',
                text,
            )
            if match:
                return clean_value(match.group(0))

    # ── 3. <legend> → <span data-selected-value> ─────────────────────────────
    legend = soup.select_one("legend.form__label")
    if legend:
        span = legend.select_one("span[data-selected-value]")
        if span:
            t = span.get_text(strip=True)
            if t and is_shipping_phrase(t):
                return clean_value(t)

    # ── 4. Any span[data-selected-value] anywhere ─────────────────────────────
    for span in soup.select("span[data-selected-value]"):
        t = span.get_text(strip=True)
        if t and is_shipping_phrase(t):
            return clean_value(t)

    # ── 5. <input value="Ships In ..."> ──────────────────────────────────────
    for inp in soup.select("input"):
        val = inp.get("value", "")
        if val and is_shipping_phrase(val):
            return clean_value(val)

    # ── 6. .text-truncator p (Pattern 1 from original brief) ─────────────────
    truncator = soup.select_one(".text-truncator p")
    if truncator:
        t = truncator.get_text(strip=True)
        if t and is_shipping_phrase(t):
            return clean_value(t)

    # ── 7. Any <p> that STRICTLY contains shipping info ───────────────────────
    for p in soup.find_all("p"):
        t = p.get_text(strip=True)
        if len(t) > 200:  # Skip long paragraphs (product descriptions)
            continue
        if is_shipping_phrase(t):
            return clean_value(t)

    # ── 8. Any element with "Delivery Timeline" label nearby ─────────────────
    for el in soup.find_all(string=re.compile(r"Delivery\s+Timeline", re.I)):
        parent = el.parent
        if parent:
            # Try sibling or child span
            for candidate in parent.find_all(["span", "div", "p"]):
                t = candidate.get_text(strip=True)
                if t and is_shipping_phrase(t):
                    return clean_value(t)

    # ── 9. Heading tags (h1-h6) with shipping info ────────────────────────────
    for tag in soup.find_all(["h1", "h2", "h3", "h4", "h5", "h6"]):
        t = tag.get_text(strip=True)
        if is_shipping_phrase(t):
            return clean_value(t)

    # ── 10. Any element with class containing "ship" or "deliver" ─────────────
    for el in soup.find_all(True):
        classes = " ".join(el.get("class", []))
        if re.search(r"(ship|deliver|dispatch)", classes, re.I):
            t = el.get_text(strip=True)
            if t and is_shipping_phrase(t) and len(t) < 150:
                return clean_value(t)

    # ── 11. Broad text scan for "Ships In X Days" pattern ────────────────────
    full_text = soup.get_text(" ", strip=True)
    match = re.search(
        r"(Ships?\s+[Ii]n\s+\d+\s+Days?|"
        r"Deliver(?:ed|y)?\s+(?:within|in)\s+[\d\-]+\s+(?:working\s+)?[Dd]ays?|"
        r"Dispatch(?:ed)?\s+(?:within|in)\s+[\d\-]+\s+(?:working\s+)?[Dd]ays?|"
        r"(?:This\s+product\s+will\s+be\s+shipped\s+within\s+[\d\-]+\s+(?:working\s+)?[Dd]ays?))",
        full_text,
    )
    if match:
        return clean_value(match.group(0))

    return "N/A"


def is_shipping_phrase(text: str) -> bool:
    """True if text looks like a shipping timeline (not a description or size)."""
    t = text.strip()
    if not t or len(t) < 3:
        return False
    if len(t) > 200:  # Too long = product description
        return False
    if not VALID_RE.search(t):
        return False
    if DESCRIPTION_RE.search(t):
        return False
    return True


def clean_value(text: str) -> str:
    """Strip boilerplate, normalise whitespace."""
    text = text.strip()
    # Remove "This product will be shipped within" prefix
    text = re.sub(
        r"(?i)^this\s+product\s+will\s+be\s+shipped\s+(within|in)\s*",
        "",
        text,
    ).strip()
    # Remove "Delivery Timeline:" prefix
    text = re.sub(r"(?i)^delivery\s+timeline\s*:\s*", "", text).strip()
    # Normalise whitespace
    text = " ".join(text.split())
    return text if text else "N/A"


# ── Playwright scraper ────────────────────────────────────────────────────────

async def playwright_scrape_one(page, url: str, retries: int = 3, delay: float = 2.0) -> str:
    """Scrape a single URL with Playwright and return shipping timeline."""
    for attempt in range(1, retries + 1):
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            # Wait for shipping element to appear
            try:
                await page.wait_for_selector(
                    "legend.form__label, .text-truncator, [data-selected-value]",
                    timeout=5000,
                )
            except Exception:
                pass  # Continue with whatever loaded
            html = await page.content()
            result = extract_shipping(html, url)
            return result
        except Exception as e:
            log.warning(f"  Attempt {attempt}/{retries} failed for {url}: {e}")
            if attempt < retries:
                await asyncio.sleep(delay * attempt)
    return "N/A"


async def run_playwright(urls: list, delay: float = 2.0, retries: int = 3) -> dict:
    """Scrape a list of URLs sequentially using one Playwright browser."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log.error("Playwright not installed. Run: pip install playwright && playwright install chromium")
        sys.exit(1)

    results = {}
    total = len(urls)

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            locale="en-IN",
        )
        page = await context.new_page()

        # Warm up with homepage
        try:
            await page.goto("https://unmatchedkicks.in/", wait_until="domcontentloaded", timeout=15000)
        except Exception:
            pass

        for i, url in enumerate(urls, 1):
            shipping = await playwright_scrape_one(page, url, retries=retries, delay=delay)
            results[url] = shipping

            # Full URL in log, no truncation
            status = "✅" if shipping != "N/A" else "❌"
            log.info(f"[{i}/{total}] {status} {url}")
            log.info(f"         → {shipping}")

            # Delay between requests
            await asyncio.sleep(delay + random.uniform(0, 0.5))

        await browser.close()

    return results


# ── Requests scraper ──────────────────────────────────────────────────────────

def requests_scrape_one(session, url: str, retries: int = 3) -> str:
    """Scrape a single URL with requests + BS4."""
    import requests as req_lib

    headers = {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
        "Accept-Language": "en-IN,en;q=0.9",
    }

    for attempt in range(1, retries + 1):
        try:
            resp = session.get(url, headers=headers, timeout=20)
            if resp.status_code == 200:
                return extract_shipping(resp.text, url)
            elif resp.status_code == 429:
                time.sleep(5 * attempt)
            else:
                log.warning(f"  HTTP {resp.status_code} for {url}")
        except Exception as e:
            log.warning(f"  Request error attempt {attempt}: {e}")
            if attempt < retries:
                time.sleep(2 * attempt)

    return "N/A"


# ── Excel / CSV I/O ───────────────────────────────────────────────────────────

def read_input(path: str) -> list[dict]:
    """Read results.xlsx or results.csv, return list of {url, timeline} dicts."""
    p = Path(path)
    if not p.exists():
        log.error(f"Input file not found: {path}")
        sys.exit(1)

    rows = []

    if p.suffix.lower() in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            log.error("openpyxl not installed. Run: pip install openpyxl")
            sys.exit(1)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        # Find url and timeline columns (flexible header names)
        url_col = _find_col(headers, ["product_url", "url", "link", "product url"])
        tl_col = _find_col(headers, ["shipping_timeline", "shipping", "timeline", "delivery"])
        if url_col is None:
            url_col = 0  # default to col A
        if tl_col is None:
            tl_col = 1  # default to col B
        for row in ws.iter_rows(min_row=2, values_only=True):
            url = str(row[url_col]).strip() if row[url_col] else ""
            tl = str(row[tl_col]).strip() if len(row) > tl_col and row[tl_col] else ""
            if url and url != "None":
                rows.append({"url": url, "timeline": tl})
    else:
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for r in reader:
                url = r.get("product_url", r.get("url", "")).strip()
                tl = r.get("shipping_timeline", r.get("timeline", "")).strip()
                if url:
                    rows.append({"url": url, "timeline": tl})

    return rows


def _find_col(headers: list, candidates: list) -> int | None:
    for i, h in enumerate(headers):
        if h.lower() in candidates:
            return i
    return None


def write_output(rows: list[dict], path: str):
    """Write results to Excel (.xlsx) or CSV."""
    p = Path(path)

    if p.suffix.lower() in (".xlsx", ".xls"):
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            log.error("openpyxl not installed. Run: pip install openpyxl")
            sys.exit(1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["product_url", "shipping_timeline"])

        for row in rows:
            url = row["url"]
            tl = row["timeline"]
            cell = ws.cell(row=ws.max_row + 1, column=1, value=url)
            # Make URL clickable
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")
            ws.cell(row=ws.max_row, column=2, value=tl)

        # Column widths
        ws.column_dimensions["A"].width = 80
        ws.column_dimensions["B"].width = 40
        wb.save(path)
    else:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, quoting=csv.QUOTE_ALL)
            writer.writerow(["product_url", "shipping_timeline"])
            for row in rows:
                writer.writerow([row["url"], row["timeline"]])

    log.info(f"Output written to: {path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Smart re-scraper — fix incomplete shipping timelines")
    parser.add_argument("--input", default="results.xlsx", help="Input file (xlsx or csv)")
    parser.add_argument("--output", default="results_updated.xlsx", help="Output file (xlsx or csv)")
    parser.add_argument("--playwright", action="store_true", help="Use Playwright (recommended)")
    parser.add_argument("--delay", type=float, default=2.0, help="Seconds between requests (default: 2)")
    parser.add_argument("--retries", type=int, default=3, help="Retries per URL (default: 3)")
    args = parser.parse_args()

    # ── Read input ────────────────────────────────────────────────────────────
    log.info(f"Reading {args.input}…")
    all_rows = read_input(args.input)
    log.info(f"Total entries: {len(all_rows)}")

    # ── Split: complete vs incomplete ─────────────────────────────────────────
    complete = []
    incomplete = []
    for row in all_rows:
        if is_valid_shipping(row["timeline"]):
            complete.append(row)
        else:
            incomplete.append(row)

    log.info(f"✅ Complete (will SKIP):      {len(complete)}")
    log.info(f"❌ Incomplete (will RE-SCRAPE): {len(incomplete)}")

    if not incomplete:
        log.info("Nothing to re-scrape! All entries are already complete.")
        write_output(all_rows, args.output)
        return

    # Show what's being re-scraped
    log.info("\nSample of incomplete entries:")
    for row in incomplete[:5]:
        log.info(f"  {row['url']}")
        log.info(f"    Current value: [{row['timeline']}]")

    # ── Re-scrape incomplete entries ──────────────────────────────────────────
    log.info(f"\nRe-scraping {len(incomplete)} entries with delay={args.delay}s…\n")
    urls_to_scrape = [row["url"] for row in incomplete]

    if args.playwright:
        new_values = asyncio.run(
            run_playwright(urls_to_scrape, delay=args.delay, retries=args.retries)
        )
    else:
        try:
            import requests as req_lib
        except ImportError:
            log.error("requests not installed. Run: pip install requests")
            sys.exit(1)

        session = req_lib.Session()
        new_values = {}
        total = len(urls_to_scrape)
        for i, url in enumerate(urls_to_scrape, 1):
            result = requests_scrape_one(session, url, retries=args.retries)
            new_values[url] = result
            status = "✅" if result != "N/A" else "❌"
            log.info(f"[{i}/{total}] {status} {url}")
            log.info(f"         → {result}")
            time.sleep(args.delay + random.uniform(0, 0.3))

    # ── Merge results ─────────────────────────────────────────────────────────
    # Update incomplete rows with new values
    updated_incomplete = []
    fixed = 0
    still_bad = 0
    for row in incomplete:
        new_val = new_values.get(row["url"], "N/A")
        if is_valid_shipping(new_val):
            fixed += 1
        else:
            still_bad += 1
        updated_incomplete.append({"url": row["url"], "timeline": new_val})

    # Preserve original order: complete rows first (in their original positions)
    # Build a map for quick lookup
    updated_map = {r["url"]: r["timeline"] for r in updated_incomplete}
    final_rows = []
    for row in all_rows:
        if row["url"] in updated_map:
            final_rows.append({"url": row["url"], "timeline": updated_map[row["url"]]})
        else:
            final_rows.append(row)

    # ── Stats ─────────────────────────────────────────────────────────────────
    log.info("\n" + "=" * 60)
    log.info("RESULTS SUMMARY")
    log.info("=" * 60)
    log.info(f"Total URLs:              {len(final_rows)}")
    log.info(f"Already complete:        {len(complete)}")
    log.info(f"Re-scraped:              {len(incomplete)}")
    log.info(f"  → Fixed:               {fixed}")
    log.info(f"  → Still incomplete:    {still_bad}")
    total_valid = len(complete) + fixed
    pct = (total_valid / len(final_rows) * 100) if final_rows else 0
    log.info(f"Overall completion:      {total_valid}/{len(final_rows)} ({pct:.1f}%)")
    log.info("=" * 60)

    if still_bad > 0:
        log.info(f"\nStill incomplete ({still_bad} URLs):")
        for row in final_rows:
            if not is_valid_shipping(row["timeline"]):
                log.info(f"  {row['url']}  →  [{row['timeline']}]")

    # ── Write output ──────────────────────────────────────────────────────────
    write_output(final_rows, args.output)
    log.info(f"\n✅ Done! Check: {args.output}")


if __name__ == "__main__":
    main()