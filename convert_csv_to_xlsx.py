import sys
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def csv_to_xlsx(csv_path: str, xlsx_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        for r_idx, row in enumerate(reader, start=1):
            for c_idx, cell in enumerate(row, start=1):
                if r_idx == 1:
                    # header
                    ws.cell(row=r_idx, column=c_idx, value=cell)
                    ws.cell(row=r_idx, column=c_idx).font = Font(bold=True)
                else:
                    if c_idx == 1:
                        # product_url column: make hyperlink
                        url = cell
                        ws.cell(row=r_idx, column=c_idx, value=url)
                        try:
                            ws.cell(row=r_idx, column=c_idx).hyperlink = url
                            ws.cell(row=r_idx, column=c_idx).style = "Hyperlink"
                        except Exception:
                            pass
                    else:
                        ws.cell(row=r_idx, column=c_idx, value=cell)

    # Auto-width columns a bit
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(val))
            except Exception:
                pass
        adjusted_width = min(max(10, max_length + 2), 80)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(xlsx_path)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python convert_csv_to_xlsx.py results.csv results.xlsx")
        sys.exit(1)
    csv_to_xlsx(sys.argv[1], sys.argv[2])
    print(f"Wrote {sys.argv[2]}")
