"""
Validation & Re-scraper for Unmatched Kicks Shipping Timeline
=============================================================

This script:
1. Reads results.xlsx (output from the main scraper)
2. Validates column B (shipping_timeline) for completeness
3. Identifies incomplete entries (e.g., "XL", "S", "M" — size data instead of shipping)
4. Re-scrapes those URLs with IMPROVED extraction patterns
5. Writes corrected results back to results.xlsx

USAGE
-----
    python validate_and_rescrape_unmatchedkicks.py --input results.xlsx --output results_fixed.xlsx

WHAT IT DETECTS AS "INCOMPLETE"
-------------------------------
✗ Empty cells
✗ Size indicators (XS, S, M, L, XL, XXL)
✗ Price data
✗ "N/A" entries
✓ Valid entries: "Ships In X Days", "This product will be shipped within...", "Same day delivery", etc.
"""

import argparse
import asyncio
import logging
import random
import re
import sys
import time
from pathlib import Path
from typing import Optional, List, Tuple, Dict
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    print("openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("validator")

# ── Constants ─────────────────────────────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_3) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
]

BASE_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-IN,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
}

# Keywords that indicate INVALID/INCOMPLETE data
INVALID_KEYWORDS = {
    r"^(xxs|xs|s|m|l|xl|xxl|one\s*size)$",  # Size indicators
    r"^\d+\.\d{2}$",  # Price (e.g., "8999.00")
    r"^(₹|rs\.?)\s*\d+",  # Indian Rupees
    r"^n/a$",  # Placeholder
    r"^(extra\s*small|small|medium|large|extra\s*large)$",  # Size words
}

# Keywords that indicate VALID shipping data
VALID_KEYWORDS = {
    r"ship",  # "Ships In", "shipped within"
    r"deliver",  # "Delivery timeline"
    r"days?",  # "Days", "day"
    r"working",  # "Working days"
    r"same\s*day",  # "Same day delivery"
    r"hours?",  # "Hours"
    r"week",  # "Week(s)"
}


def is_valid_shipping_text(text: str) -> bool:
    """Check if text is valid shipping timeline data."""
    text_lower = text.lower().strip()
    
    # Check for invalid patterns (sizes, prices, etc.)
    for pattern in INVALID_KEYWORDS:
        if re.search(pattern, text_lower, re.IGNORECASE):
            return False
    
    # Check for valid patterns
    for pattern in VALID_KEYWORDS:
        if re.search(pattern, text_lower, re.IGNORECASE):
            return True
    
    return False


def validate_shipping_cell(value: str) -> bool:
    """Return True if the cell contains valid shipping timeline data."""
    if not value or not str(value).strip():
        return False
    
    value_str = str(value).strip()
    return is_valid_shipping_text(value_str)


def read_excel_for_validation(input_path: str) -> List[Tuple[int, str, str]]:
    """
    Read results.xlsx and return list of (row_number, product_url, shipping_timeline) tuples.
    Filter for incomplete entries.
    """
    try:
        wb = load_workbook(input_path)
        ws = wb.active
    except Exception as e:
        log.error(f"Failed to load {input_path}: {e}")
        return []
    
    incomplete = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        url_cell = row[0]
        timeline_cell = row[1]
        
        url = url_cell.value
        timeline = timeline_cell.value
        
        if not url:
            continue
        
        # Check if this entry is incomplete
        if not validate_shipping_cell(timeline):
            incomplete.append((row_idx, str(url), str(timeline) if timeline else ""))
    
    return incomplete


# ── IMPROVED Extraction Logic ────────────────────────────────────────────────

def extract_shipping_v2(html: str) -> str:
    """
    ENHANCED extraction with additional patterns.
    Priority:
    1. Structured data (JSON-LD, script tags)
    2. Button/form element with "SHIPS IN X DAYS"
    3. Delivery timeline section (heading + text)
    4. Product description paragraphs
    5. Any span/div with shipping keywords
    """
    soup = BeautifulSoup(html, "html.parser")
    
    # ── Pattern 0: JSON-LD Structured Data ──────────────────────────────────
    scripts = soup.find_all("script", {"type": "application/ld+json"})
    for script in scripts:
        try:
            import json
            data = json.loads(script.string)
            if isinstance(data, dict):
                # Look for aggregateOffer > offers > shippingDetails
                if "aggregateOffer" in data:
                    agg = data["aggregateOffer"]
                    if isinstance(agg, dict) and "offers" in agg:
                        offers = agg["offers"]
                        if isinstance(offers, list) and len(offers) > 0:
                            offer = offers[0]
                            if isinstance(offer, dict) and "shippingDetails" in offer:
                                shipping = offer["shippingDetails"]
                                if isinstance(shipping, dict) and "deliveryTime" in shipping:
                                    delivery = shipping["deliveryTime"]
                                    if isinstance(delivery, dict) and "businessDays" in delivery:
                                        days = delivery["businessDays"]
                                        return clean_text(f"Ships In {days} Days")
        except:
            pass
    
    # ── Pattern 1a: Button with "SHIPS IN X DAYS" ───────────────────────────
    button = soup.find("button", string=re.compile(r"ships?\s+in", re.I))
    if button:
        text = button.get_text(strip=True)
        if text:
            return clean_text(text)
    
    # ── Pattern 1b: Any element with text-content "SHIPS IN X DAYS" ─────────
    for tag in soup.find_all(string=re.compile(r"ships?\s+in\s+\d+\s+days?", re.I)):
        text = tag.strip()
        if text:
            return clean_text(text)
    
    # ── Pattern 2: Legend + Span (form-based delivery selector) ─────────────
    span = soup.select_one("legend.form__label span[data-selected-value]")
    if span:
        text = span.get_text(strip=True)
        if text and is_valid_shipping_text(text):
            return clean_text(text)
    
    # ── Pattern 3: Input value attribute ────────────────────────────────────
    inp = soup.select_one('input[value*="Ships"], input[value*="ships"]')
    if inp:
        text = inp.get("value", "").strip()
        if text and is_valid_shipping_text(text):
            return clean_text(text)
    
    # ── Pattern 4: Text-truncator paragraph (description with shipping info) ─
    p = soup.select_one(".text-truncator p")
    if p:
        text = p.get_text(strip=True)
        if text and is_valid_shipping_text(text):
            return clean_text(text)
    
    # ── Pattern 5: Heading containing "DELIVERY TIMELINE" + sibling text ────
    for heading in soup.find_all(["h2", "h3", "h4", "div"], string=re.compile(r"delivery\s+timeline", re.I)):
        # Look for sibling with shipping info
        for sibling in heading.find_next_siblings():
            text = sibling.get_text(strip=True)
            if text and is_valid_shipping_text(text):
                return clean_text(text)
    
    # ── Pattern 6: Badge/Label with shipping keywords ──────────────────────
    for badge in soup.find_all(["span", "div"], class_=re.compile(r"badge|label|tag", re.I)):
        text = badge.get_text(strip=True)
        if text and is_valid_shipping_text(text) and len(text) < 100:
            return clean_text(text)
    
    # ── Pattern 7: Broad search - any paragraph mentioning shipping ────────
    for p_tag in soup.find_all("p"):
        text = p_tag.get_text(strip=True)
        if text and is_valid_shipping_text(text) and len(text) < 200:
            return clean_text(text)
    
    # ── Pattern 8: Fallback - check all divs with text nodes ───────────────
    for div in soup.find_all("div", limit=500):  # reasonable limit
        text = div.get_text(strip=True)
        if text and is_valid_shipping_text(text) and 5 < len(text) < 150:
            return clean_text(text)
    
    return "N/A"


def clean_text(text: str) -> str:
    """Clean boilerplate and normalize whitespace."""
    text = text.strip()
    # Remove common prefixes
    text = re.sub(r"(?i)^delivery\s+timeline\s*:\s*", "", text).strip()
    text = re.sub(r"(?i)^this product will be shipped (within|in)\s*", "", text).strip()
    text = re.sub(r"(?i)^shipped (within|in)\s*", "", text).strip()
    text = re.sub(r"(?i)^ships?\s+(in|within)\s+", "Ships In ", text).strip()
    # Normalize internal whitespace
    text = " ".join(text.split())
    return text if text else "N/A"


# ── HTTP Scraper (requests) ───────────────────────────────────────────────────

class RequestsScraper:
    def __init__(self, retries: int = 3, delay: float = 2.0):
        self.retries = retries
        self.delay = delay
        self.session = requests.Session()
        try:
            self.session.get(
                "https://unmatchedkicks.in/",
                headers=self._headers(),
                timeout=15,
            )
        except Exception:
            pass

    def _headers(self) -> dict:
        h = dict(BASE_HEADERS)
        h["User-Agent"] = random.choice(USER_AGENTS)
        return h

    def fetch(self, url: str) -> Optional[str]:
        for attempt in range(1, self.retries + 1):
            try:
                resp = self.session.get(
                    url,
                    headers=self._headers(),
                    timeout=20,
                    allow_redirects=True,
                )
                if resp.status_code == 200:
                    return resp.text
                elif resp.status_code in (429, 503):
                    wait = self.delay * (2 ** attempt)
                    log.warning(f"Rate limited ({resp.status_code}) on {url}. Waiting {wait:.1f}s…")
                    time.sleep(wait)
                elif resp.status_code == 404:
                    log.warning(f"404 Not Found: {url}")
                    return None
                else:
                    log.warning(f"HTTP {resp.status_code} for {url} (attempt {attempt})")
            except requests.exceptions.RequestException as e:
                log.warning(f"Request error on {url} attempt {attempt}: {e}")
                if attempt < self.retries:
                    time.sleep(self.delay * attempt)
        return None

    def scrape(self, url: str) -> Optional[str]:
        html = self.fetch(url)
        if html is None:
            return None
        return extract_shipping_v2(html)


# ── Playwright Scraper (async) ────────────────────────────────────────────────

async def playwright_scrape_single(url: str, delay: float = 2.0, retries: int = 3) -> Optional[str]:
    """Scrape a single URL with Playwright."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log.error("Playwright not installed. Install with: pip install playwright && playwright install chromium")
        return None

    for attempt in range(1, retries + 1):
        try:
            async with async_playwright() as pw:
                browser = await pw.chromium.launch(headless=True)
                context = await browser.new_context(
                    user_agent=random.choice(USER_AGENTS),
                    locale="en-IN",
                    extra_http_headers={"Accept-Language": "en-IN,en;q=0.9"},
                )
                page = await context.new_page()
                
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                
                # Wait for shipping-related elements
                try:
                    await page.wait_for_selector(
                        "legend.form__label, .text-truncator, button:has-text('SHIPS'), [class*='delivery'], [class*='timeline']",
                        timeout=5000,
                    )
                except:
                    pass
                
                await asyncio.sleep(0.5)
                html = await page.content()
                await browser.close()
                
                shipping = extract_shipping_v2(html)
                return shipping if shipping != "N/A" else None
        except Exception as e:
            log.warning(f"[PW] Error on {url} attempt {attempt}: {e}")
            if attempt == retries:
                return None
            await asyncio.sleep(delay * attempt)
    
    return None


# ── Main Workflow ────────────────────────────────────────────────────────────

def validate_and_rescrape(
    input_path: str,
    output_path: str,
    use_playwright: bool = False,
    workers: int = 2,
    delay: float = 2.0,
):
    """Main validation and re-scraping workflow."""
    log.info(f"Reading {input_path}…")
    incomplete_entries = read_excel_for_validation(input_path)
    
    if not incomplete_entries:
        log.info("✅ All entries are valid! No re-scraping needed.")
        return
    
    log.info(f"Found {len(incomplete_entries)} incomplete entries to re-scrape")
    log.info(f"Using {'Playwright' if use_playwright else 'Requests'} mode")
    
    scraper = RequestsScraper(retries=3, delay=delay)
    fixed_results = {}
    
    if use_playwright:
        # Batch Playwright requests
        urls_to_scrape = [url for _, url, _ in incomplete_entries]
        results = asyncio.run(playwright_batch_scrape(urls_to_scrape, delay=delay))
        for entry in incomplete_entries:
            row_idx, url, _ = entry
            fixed_results[row_idx] = results.get(url, "N/A")
    else:
        # Threaded requests
        def process(entry):
            row_idx, url, old_value = entry
            log.info(f"Re-scraping: {url} (was: {old_value})")
            time.sleep(delay + random.uniform(0, 1))
            result = scraper.scrape(url)
            return row_idx, result if result else "N/A"
        
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {pool.submit(process, entry): entry for entry in incomplete_entries}
            for future in as_completed(futures):
                row_idx, new_value = future.result()
                fixed_results[row_idx] = new_value
                log.info(f"  → Result: {new_value}")
    
    # Write corrected results back to Excel
    log.info(f"\nWriting corrected results to {output_path}…")
    write_corrected_excel(input_path, output_path, fixed_results)
    log.info("✅ Done!")


async def playwright_batch_scrape(urls: List[str], delay: float = 2.0) -> Dict[str, str]:
    """Scrape multiple URLs with Playwright in batch."""
    results = {}
    for url in urls:
        result = await playwright_scrape_single(url, delay=delay)
        results[url] = result if result else "N/A"
        log.info(f"[PW] {url} → {results[url]}")
        await asyncio.sleep(delay)
    return results


def write_corrected_excel(input_path: str, output_path: str, fixed_results: Dict[int, str]):
    """Write corrected data back to Excel."""
    wb = load_workbook(input_path)
    ws = wb.active
    
    for row_idx, new_value in fixed_results.items():
        ws.cell(row=row_idx, column=2, value=new_value)
        # Mark corrected cells with a light green background
        if new_value != "N/A":
            cell = ws.cell(row=row_idx, column=2)
            cell.fill = None  # Could add color here with PatternFill
    
    wb.save(output_path)
    log.info(f"Saved {len(fixed_results)} corrected entries")


# ── CLI ───────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Validate and re-scrape incomplete shipping timelines",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--input", default="results.xlsx", help="Input Excel file (default: results.xlsx)")
    p.add_argument("--output", default="results_fixed.xlsx", help="Output Excel file (default: results_fixed.xlsx)")
    p.add_argument("--playwright", action="store_true", help="Use Playwright for re-scraping")
    p.add_argument("--workers", type=int, default=2, help="Concurrent workers (default: 2)")
    p.add_argument("--delay", type=float, default=2.0, help="Delay between requests (default: 2.0)")
    return p


def main():
    parser = build_parser()
    args = parser.parse_args()
    
    if not Path(args.input).exists():
        log.error(f"Input file not found: {args.input}")
        sys.exit(1)
    
    validate_and_rescrape(
        input_path=args.input,
        output_path=args.output,
        use_playwright=args.playwright,
        workers=args.workers,
        delay=args.delay,
    )


if __name__ == "__main__":
    main()