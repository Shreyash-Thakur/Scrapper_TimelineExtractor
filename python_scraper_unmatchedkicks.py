"""
Shipping Timeline Scraper — unmatchedkicks.in (FINAL VERSION)
==============================================================
Run this script on YOUR LOCAL MACHINE (not a cloud server).
The site blocks datacenter IPs; residential/home IPs work fine.

FEATURES
--------
✓ Playwright mode for JavaScript-rendered content (recommended)
✓ Requests mode with session management
✓ Intelligent detection of shipping data vs. size data
✓ Multiple extraction patterns (JSON-LD, buttons, form elements, text content)
✓ Excel output with hyperlinked URLs
✓ Detailed logging and retry logic

USAGE
-----
# Install dependencies (once):
    pip install requests beautifulsoup4 playwright openpyxl
    playwright install chromium

# Recommended (Playwright - handles JavaScript):
    python python_scraper_unmatchedkicks.py --input urls.txt --output results.xlsx --playwright --delay 3

# Alternative (Requests - faster but less reliable):
    python python_scraper_unmatchedkicks.py --input urls.txt --output results.xlsx --workers 2 --delay 2.5

# Re-validate and fix incomplete entries:
    python validate_and_rescrape_unmatchedkicks.py --input results.xlsx --output results_fixed.xlsx --playwright

INPUT FILE FORMAT (urls.txt)
-----------------------------
One URL per line. Lines starting with # or blank lines are skipped.
Example:
    https://unmatchedkicks.in/products/sp5der-sp555-tee-black
    https://unmatchedkicks.in/products/hellstar-warm-up-t-shirt-black

OUTPUT
------
  results.xlsx      — product_url (clickable links), shipping_timeline
  failed_urls.txt   — URLs that failed after all retries
"""

import argparse
import asyncio
import logging
import random
import re
import sys
import time
from pathlib import Path
from typing import Optional, List
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("scraper")

# ── Constants ─────────────────────────────────────────────────────────────────
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:123.0) Gecko/20100101 Firefox/123.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_3) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.3 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
]

BASE_HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-IN,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
}


# ── Validation & Extraction Logic ────────────────────────────────────────────

VALID_SHIPPING_KEYWORDS = {
    r"ship",  # "Ships In", "shipped within"
    r"deliver",  # "Delivery timeline", "deliver"
    r"days?",  # "Days", "day"
    r"working",  # "Working days"
    r"same\s*day",  # "Same day delivery"
    r"hours?",  # "Hours"
    r"week",  # "Week(s)"
}

INVALID_KEYWORDS = {
    r"^(xxs|xs|s|m|l|xl|xxl|one\s*size)$",  # Size indicators
    r"^\d+\.\d{2}$",  # Price (e.g., "8999.00")
    r"^(₹|rs\.?)\s*\d+",  # Indian Rupees
    r"^n/a$",  # Placeholder
}


def is_valid_shipping_text(text: str) -> bool:
    """Check if text is valid shipping timeline data."""
    text_lower = text.lower().strip()
    
    # Check for invalid patterns (sizes, prices, etc.)
    for pattern in INVALID_KEYWORDS:
        if re.search(pattern, text_lower, re.IGNORECASE):
            return False
    
    # Check for valid patterns
    for pattern in VALID_SHIPPING_KEYWORDS:
        if re.search(pattern, text_lower, re.IGNORECASE):
            return True
    
    return False


def extract_shipping(html: str) -> str:
    """
    ENHANCED extraction with multiple patterns and validation.
    Priority:
    1. JSON-LD structured data
    2. Button/form with "SHIPS IN X DAYS"
    3. Legend + span (form-based selector)
    4. Input value attribute
    5. Text-truncator paragraph
    6. Delivery timeline heading + content
    7. Badges/labels with shipping keywords
    8. General paragraphs with shipping info
    """
    soup = BeautifulSoup(html, "html.parser")
    
    # ── Pattern 0: JSON-LD Structured Data ──────────────────────────────────
    scripts = soup.find_all("script", {"type": "application/ld+json"})
    for script in scripts:
        try:
            import json
            data = json.loads(script.string)
            if isinstance(data, dict) and "aggregateOffer" in data:
                agg = data["aggregateOffer"]
                if isinstance(agg, dict) and "offers" in agg:
                    offers = agg["offers"]
                    if isinstance(offers, list) and len(offers) > 0:
                        offer = offers[0]
                        if isinstance(offer, dict) and "shippingDetails" in offer:
                            shipping = offer["shippingDetails"]
                            if isinstance(shipping, dict) and "deliveryTime" in shipping:
                                delivery = shipping["deliveryTime"]
                                if isinstance(delivery, dict) and "businessDays" in delivery:
                                    days = delivery["businessDays"]
                                    text = f"Ships In {days} Days"
                                    if is_valid_shipping_text(text):
                                        return clean_text(text)
        except:
            pass
    
    # ── Pattern 1: Button with "SHIPS IN X DAYS" ───────────────────────────
    for button in soup.find_all(["button", "div"], string=re.compile(r"ships?\s+in", re.I)):
        text = button.get_text(strip=True)
        if text and is_valid_shipping_text(text):
            return clean_text(text)
    
    # ── Pattern 2: Legend + Span (form-based delivery selector) ─────────────
    span = soup.select_one("legend.form__label span[data-selected-value]")
    if span:
        text = span.get_text(strip=True)
        if text and is_valid_shipping_text(text):
            return clean_text(text)
    
    # ── Pattern 3: Input value attribute ────────────────────────────────────
    inp = soup.select_one('input[value*="Ships"], input[value*="ships"]')
    if inp:
        text = inp.get("value", "").strip()
        if text and is_valid_shipping_text(text):
            return clean_text(text)
    
    # ── Pattern 4: Text-truncator paragraph ──────────────────────────────────
    p = soup.select_one(".text-truncator p")
    if p:
        text = p.get_text(strip=True)
        if text and is_valid_shipping_text(text):
            return clean_text(text)
    
    # ── Pattern 5: Delivery timeline heading + sibling ──────────────────────
    for heading in soup.find_all(["h2", "h3", "h4", "div"], string=re.compile(r"delivery\s+timeline", re.I)):
        for sibling in heading.find_next_siblings(["div", "p", "span"], limit=3):
            text = sibling.get_text(strip=True)
            if text and is_valid_shipping_text(text):
                return clean_text(text)
    
    # ── Pattern 6: Badge/Label with shipping keywords ──────────────────────
    for badge in soup.find_all(["span", "div"], class_=re.compile(r"badge|label|tag|pill", re.I)):
        text = badge.get_text(strip=True)
        if text and is_valid_shipping_text(text) and 5 < len(text) < 100:
            return clean_text(text)
    
    # ── Pattern 7: Blue highlighted box (common pattern) ────────────────────
    for div in soup.find_all("div", style=re.compile(r"background", re.I)):
        text = div.get_text(strip=True)
        if text and is_valid_shipping_text(text) and 5 < len(text) < 150:
            return clean_text(text)
    
    # ── Pattern 8: Broad search - paragraphs with shipping keywords ────────
    for p_tag in soup.find_all("p", limit=100):
        text = p_tag.get_text(strip=True)
        if text and is_valid_shipping_text(text) and 5 < len(text) < 200:
            return clean_text(text)
    
    return "N/A"


def clean_text(text: str) -> str:
    """Remove boilerplate prefixes and normalize whitespace."""
    text = text.strip()
    # Remove common prefixes
    text = re.sub(r"(?i)^delivery\s+timeline\s*:\s*", "", text).strip()
    text = re.sub(r"(?i)^this product will be shipped (within|in)\s*", "", text).strip()
    text = re.sub(r"(?i)^shipped (within|in)\s*", "", text).strip()
    text = re.sub(r"(?i)^ships?\s+(in|within)\s+", "Ships In ", text).strip()
    # Normalize whitespace
    text = " ".join(text.split())
    return text if text else "N/A"


# ── HTTP Scraper (requests) ───────────────────────────────────────────────────

class RequestsScraper:
    def __init__(self, retries: int = 3, delay: float = 2.0):
        self.retries = retries
        self.delay = delay
        self.session = requests.Session()
        try:
            self.session.get(
                "https://unmatchedkicks.in/",
                headers=self._headers(),
                timeout=15,
            )
        except Exception:
            pass

    def _headers(self) -> dict:
        h = dict(BASE_HEADERS)
        h["User-Agent"] = random.choice(USER_AGENTS)
        return h

    def fetch(self, url: str) -> Optional[str]:
        for attempt in range(1, self.retries + 1):
            try:
                resp = self.session.get(
                    url,
                    headers=self._headers(),
                    timeout=20,
                    allow_redirects=True,
                )
                if resp.status_code == 200:
                    return resp.text
                elif resp.status_code in (429, 503):
                    wait = self.delay * (2 ** attempt)
                    log.warning(f"Rate limited ({resp.status_code}) on {url}. Waiting {wait:.1f}s…")
                    time.sleep(wait)
                elif resp.status_code == 404:
                    log.warning(f"404 Not Found: {url}")
                    return None
                else:
                    log.warning(f"HTTP {resp.status_code} for {url}")
            except requests.exceptions.RequestException as e:
                log.warning(f"Request error on {url} attempt {attempt}: {e}")
                if attempt < self.retries:
                    time.sleep(self.delay * attempt)
        return None

    def scrape(self, url: str) -> Optional[str]:
        html = self.fetch(url)
        if html is None:
            return None
        return extract_shipping(html)


# ── Playwright Scraper (async) ────────────────────────────────────────────────

async def playwright_scrape_batch(urls: List[str], delay: float = 2.0, retries: int = 3):
    """Scrape a list of URLs using Playwright."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log.error("Playwright not installed. Run: pip install playwright && playwright install chromium")
        sys.exit(1)

    results = {}
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            locale="en-IN",
            extra_http_headers={"Accept-Language": "en-IN,en;q=0.9"},
        )
        page = await context.new_page()

        for url in urls:
            for attempt in range(1, retries + 1):
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                    
                    # Wait for shipping-related elements
                    try:
                        await page.wait_for_selector(
                            "legend.form__label, .text-truncator, button, [class*='delivery']",
                            timeout=5000,
                        )
                    except:
                        pass
                    
                    await asyncio.sleep(0.5)
                    html = await page.content()
                    shipping = extract_shipping(html)
                    results[url] = shipping
                    log.info(f"[PW] {url} → {shipping}")
                    await asyncio.sleep(delay + random.uniform(0, 0.5))
                    break
                except Exception as e:
                    log.warning(f"[PW] Error on {url} attempt {attempt}: {e}")
                    if attempt == retries:
                        results[url] = "N/A"
                    else:
                        await asyncio.sleep(delay * attempt)

        await browser.close()
    return results


# ── Main Orchestration ────────────────────────────────────────────────────────

def load_urls(path: str) -> List[str]:
    """Load URLs from text file."""
    urls = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                if not line.startswith("http"):
                    line = f"https://unmatchedkicks.in/products/{line}"
                urls.append(line)
    return urls


def run_requests_mode(
    urls: List[str],
    output_path: str,
    failed_path: str,
    workers: int,
    delay: float,
    retries: int,
):
    """Run scraper using requests + threading."""
    scraper = RequestsScraper(retries=retries, delay=delay)
    results = []
    failed = []
    total = len(urls)

    log.info(f"Starting requests mode | {total} URLs | {workers} workers | delay={delay}s")

    def process(url: str):
        time.sleep(delay + random.uniform(0, 1))
        shipping = scraper.scrape(url)
        return url, shipping

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(process, url): url for url in urls}
        done = 0
        for future in as_completed(futures):
            url, shipping = future.result()
            done += 1
            if shipping is None:
                failed.append(url)
                log.error(f"[{done}/{total}] FAILED: {url}")
            else:
                results.append({"product_url": url, "shipping_timeline": shipping})
                log.info(f"[{done}/{total}] {url} → {shipping}")
            time.sleep(delay / workers)

    write_results_excel(results, output_path)
    write_failed(failed, failed_path)
    log.info(f"\n✅ Done. {len(results)} scraped | {len(failed)} failed")
    log.info(f"   Output:  {output_path}")
    log.info(f"   Failed:  {failed_path}")


def run_playwright_mode(
    urls: List[str],
    output_path: str,
    failed_path: str,
    delay: float,
    retries: int,
):
    """Run scraper using Playwright."""
    log.info(f"Starting Playwright mode | {len(urls)} URLs | delay={delay}s")
    results_map = asyncio.run(playwright_scrape_batch(urls, delay=delay, retries=retries))

    results = []
    failed = []
    for url, shipping in results_map.items():
        if shipping is None or shipping == "N/A":
            failed.append(url)
        else:
            results.append({"product_url": url, "shipping_timeline": shipping})

    write_results_excel(results, output_path)
    write_failed(failed, failed_path)
    log.info(f"\n✅ Done. {len(results)} scraped | {len(failed)} failed")
    log.info(f"   Output:  {output_path}")
    log.info(f"   Failed:  {failed_path}")


def write_results_excel(rows: List[dict], path: str):
    """Write results to Excel with hyperlinked URLs."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"

    # Headers
    headers = ["product_url", "shipping_timeline"]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        url = row["product_url"]
        timeline = row["shipping_timeline"]
        
        # Hyperlink in column A
        url_cell = sheet.cell(row=row_idx, column=1, value=url)
        url_cell.hyperlink = url
        url_cell.font = Font(color="0563C1", underline="single")
        
        # Timeline in column B
        sheet.cell(row=row_idx, column=2, value=timeline)

    # Column widths
    sheet.column_dimensions["A"].width = 75
    sheet.column_dimensions["B"].width = 30

    workbook.save(path)


def write_failed(urls: List[str], path: str):
    """Write failed URLs to text file."""
    with open(path, "w", encoding="utf-8") as f:
        for url in urls:
            f.write(url + "\n")


# ── CLI ───────────────────────────────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Scrape shipping timelines from unmatchedkicks.in",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--input", default="urls.txt", help="Input file with URLs (default: urls.txt)")
    p.add_argument("--output", default="results.xlsx", help="Output Excel file (default: results.xlsx)")
    p.add_argument("--failed", default="failed_urls.txt", help="Failed URLs file (default: failed_urls.txt)")
    p.add_argument("--workers", type=int, default=2, help="Concurrent workers (default: 2)")
    p.add_argument("--delay", type=float, default=2.5, help="Delay between requests (default: 2.5)")
    p.add_argument("--retries", type=int, default=3, help="Retry attempts (default: 3)")
    p.add_argument("--playwright", action="store_true", help="Use Playwright instead of requests")
    return p


def main():
    parser = build_parser()
    args = parser.parse_args()

    try:
        urls = load_urls(args.input)
    except FileNotFoundError:
        log.error(f"Input file not found: {args.input}")
        sys.exit(2)

    if not urls:
        log.error("No URLs loaded. Check your input file.")
        sys.exit(1)

    log.info(f"Loaded {len(urls)} URLs from {args.input}")

    if args.playwright:
        run_playwright_mode(
            urls=urls,
            output_path=args.output,
            failed_path=args.failed,
            delay=args.delay,
            retries=args.retries,
        )
    else:
        run_requests_mode(
            urls=urls,
            output_path=args.output,
            failed_path=args.failed,
            workers=args.workers,
            delay=args.delay,
            retries=args.retries,
        )


if __name__ == "__main__":
    main()